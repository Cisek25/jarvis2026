"""
Generate Wdrozenia_IdoBooking.xlsx — new architecture:
  Sheet 1  "Wdrożenia IdoB"  — data rows only (rows 1-16), NOTHING below
  Sheet 2  "Dashboard"       — all stats, pipeline, goal + 5 charts
                               all formulas reference D2:D1000 etc. for auto-expansion
openpyxl 3.1.x
"""

import sys
sys.path.insert(0, "/private/tmp/claude-502/pylibs")

import os
from datetime import date
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import DoughnutChart, PieChart, BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint

OUTPUT_PATH = "/Users/user/Desktop/Claude_STRONY/claudedocs/Wdrozenia_IdoBooking.xlsx"

# ── colour palette ────────────────────────────────────────────────────────────
DK_GREEN   = "1a472a"
MD_GREEN   = "70ad47"
LT_GREEN   = "f0f7ee"
WHITE      = "FFFFFF"
LT_GRAY    = "f2f2f2"

S_ZAKONCZONA = ("c6efce", "276221")
S_W_TOKU     = ("bdd7ee", "1f4e79")
S_BRIEF      = ("ffeb9c", "9c5700")
S_POPRAWKI   = ("ffc7ce", "9c0006")

STATUS_MAP = {
    "Zakończona":     S_ZAKONCZONA,
    "W toku":         S_W_TOKU,
    "Brief":          S_BRIEF,
    "Poprawki po WD": S_POPRAWKI,
}

# ── reusable style objects ────────────────────────────────────────────────────
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

def _fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)

def _font(bold=False, color="000000", size=11, name="Calibri"):
    return Font(bold=bold, color=color, size=size, name=name)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _hdr_cell(ws, row, col, value, bg=MD_GREEN, fg=WHITE, size=11, h_align="center"):
    """Dark header cell — green bg, white bold text."""
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _font(bold=True, color=fg, size=size)
    c.fill      = _fill(bg)
    c.alignment = _align(h=h_align)
    c.border    = THIN_BORDER
    return c

def _data_cell(ws, row, col, value, bg=WHITE, fg="000000",
               bold=False, h_align="left", num_fmt=None, font_name="Calibri"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _font(bold=bold, color=fg, name=font_name)
    c.fill      = _fill(bg)
    c.alignment = _align(h=h_align)
    c.border    = THIN_BORDER
    if num_fmt:
        c.number_format = num_fmt
    return c


# ══════════════════════════════════════════════════════════════════════════════
# PROJECT DATA
# ══════════════════════════════════════════════════════════════════════════════

HEADERS = [
    "l.p.", "Nazwa firmy", "ID panelu", "Status", "Strona WWW",
    "Data startu", "Data dostarczenia", "Skomplikowanie", "Języki",
    "SLA budowy (dni rob.)", "SLA budowy (h)", "Czas rzecz. (h)",
    "Czas rzecz. (dni rob.)", "Delta SLA (h)", "SLA dotrzymane?",
    "Rundy poprawek", "Liczba zmian", "SLA poprawki (dni)",
    "Feedback klienta", "BOTTLENECK?",
]

# col index → width
COL_WIDTHS = {
    1:5, 2:35, 3:10, 4:18, 5:35, 6:14, 7:14, 8:16, 9:14,
    10:12, 11:12, 12:12, 13:13, 14:12, 15:14, 16:14, 17:13,
    18:14, 19:15, 20:16,
}

PROJECTS = [
    (1,  "WHITE FROG (CityApart Szczecin)",        56026, "Zakończona",     "https://cityapartszczecin.pl/",         date(2026,1,15), date(2026,1,25), "Średnia",       "PL",          7,  56,  8,  1, "+48h", "✓ TAK", 0,  0,  "-",     "NIE", ""),
    (2,  "WARSZAWSKI HOLDING (WawaBed)",            55761, "Zakończona",     "https://wh.waw.pl/",                    date(2026,1,20), date(2026,2,5),  "Wysoka",        "PL",          10, 80,  20, 3, "+60h", "✓ TAK", 0,  0,  "-",     "NIE", ""),
    (3,  "VILLA KAPITAŃSKA",                        57041, "Zakończona",     "https://villa-kapitanska.pl/",          date(2026,1,25), date(2026,2,15), "Bardzo wysoka", "PL+DE",       14, 112, 30, 4, "+82h", "✓ TAK", 0,  0,  "-",     "NIE", ""),
    (4,  "Mountain Prestige MAR",                   57060, "Brief",          "https://client57060.idosell.com/",      None,            None,            "Brief",         "?",           0,  0,   0,  0, "-",    "-",     0,  0,  "-",     "NIE", ""),
    (5,  "MORADA SERVICE (GoldenApartments)",        57304, "Zakończona",     "https://client57304.idosell.com/",      date(2026,2,10), date(2026,2,28), "Średnia",       "PL",          7,  56,  15, 2, "+41h", "✓ TAK", 2,  6,  "2 dni", "TAK", ""),
    (6,  "MW Consulting",                           53370, "Zakończona",     "https://client53370.idosell.com/",      date(2026,2,15), date(2026,2,25), "Średnia",       "PL",          7,  56,  10, 2, "+46h", "✓ TAK", 0,  0,  "-",     "NIE", ""),
    (7,  "GeoStay LLC",                             57156, "Zakończona",     "https://geostay.ge/",                   date(2026,2,20), date(2026,3,5),  "Wysoka",        "PL+RU",       10, 80,  14, 2, "+66h", "✓ TAK", 0,  0,  "-",     "NIE", ""),
    (8,  "WebActive (Mazurski Chill)",              57689, "W toku",         "https://client57689.idosell.com/",      date(2026,3,1),  None,            "Średnia",       "PL",          7,  56,  12, 2, "+44h", "✓ TAK", 1,  4,  "2 dni", "NIE", ""),
    (9,  "MADERA (Madera Centrum)",                 45553, "Zakończona",     "http://www.maderacentrum.pl/",          date(2026,2,1),  date(2026,3,17), "Wysoka",        "PL+EN",       10, 80,  20, 3, "+60h", "✓ TAK", 3,  18, "3 dni", "TAK", "⚠ UWAGA"),
    (10, "NAJMAR",                                  57656, "Zakończona",     "https://client57656.idosell.com/",      date(2026,1,5),  date(2026,3,30), "Bardzo wysoka", "PL",          14, 112, 35, 5, "+77h", "✓ TAK", 14, 33, "2 dni", "TAK", "🔴 BOTTLENECK"),
    (11, "EcoCamping (Krzysztof Dostal)",           57820, "Zakończona",     "http://client57820.idosell.com/",       date(2026,3,10), date(2026,4,3),  "Bardzo wysoka", "PL+EN",       14, 112, 28, 4, "+84h", "✓ TAK", 2,  22, "3 dni", "TAK", ""),
    (12, "MENTALIS",                                23326, "Brief",          "https://client23326.idosell.com/",      None,            None,            "Brief",         "?",           0,  0,   0,  0, "-",    "-",     0,  0,  "-",     "NIE", ""),
    (13, "MARCIN SKRZYPCZAK",                       33019, "Brief",          "https://client33019.idosell.com/",      None,            None,            "Brief",         "?",           0,  0,   0,  0, "-",    "-",     0,  0,  "-",     "NIE", ""),
    (14, "WCA Wrocław City Apartments",             6645,  "Poprawki po WD", "https://wroclawcityapartments.pl/",     date(2026,3,20), None,            "Niska",         "PL",          3,  24,  5,  1, "+19h", "✓ TAK", 0,  0,  "3 dni", "NIE", ""),
    (15, "SORS Eagle Tower (Paweł Pasławski)",      51651, "Poprawki po WD", "https://client51651.idosell.com/",      date(2026,3,15), None,            "Bardzo wysoka", "PL+EN+DE+ES", 14, 112, 22, 3, "+90h", "✓ TAK", 2,  9,  "2 dni", "NIE", ""),
]

# columns that get center alignment in data rows
CENTER_COLS = {1, 3, 4, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19}


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — "Wdrożenia IdoB"
# ══════════════════════════════════════════════════════════════════════════════

def build_sheet1(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Wdrożenia IdoB"

    # ── column widths ─────────────────────────────────────────────────────────
    for col_idx, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── header row 1 ──────────────────────────────────────────────────────────
    for col_idx, hdr in enumerate(HEADERS, start=1):
        _hdr_cell(ws, 1, col_idx, hdr, bg=DK_GREEN, fg=WHITE, size=11,
                  h_align="center")
    ws.row_dimensions[1].height = 30

    # ── data rows 2-16 ────────────────────────────────────────────────────────
    for offset, proj in enumerate(PROJECTS):
        row = offset + 2
        (lp, name, panel_id, status, url, start, delivered, complexity, langs,
         sla_days, sla_h, real_h, real_days, delta, sla_met,
         rounds, changes, sla_fix, feedback, bottleneck) = proj

        base_bg = LT_GRAY if row % 2 == 0 else WHITE

        row_values = [
            lp, name, panel_id, status, url, start, delivered,
            complexity, langs, sla_days, sla_h, real_h, real_days,
            delta, sla_met, rounds, changes, sla_fix, feedback, bottleneck,
        ]

        for col_idx, val in enumerate(row_values, start=1):
            h = "center" if col_idx in CENTER_COLS else "left"
            c = _data_cell(ws, row, col_idx, val, bg=base_bg, h_align=h)

            # date format
            if col_idx in (6, 7) and isinstance(val, date):
                c.number_format = "DD.MM.YYYY"
            # integer number format for numeric columns
            if col_idx in (11, 12, 13, 16, 17) and isinstance(val, int):
                c.number_format = "0"

        # ── Column D — status colouring ───────────────────────────────────────
        if status in STATUS_MAP:
            bg, fg = STATUS_MAP[status]
            d = ws.cell(row=row, column=4)
            d.fill = _fill(bg)
            d.font = _font(bold=True, color=fg)

        # ── Column O — SLA dotrzymane colouring ───────────────────────────────
        o_cell = ws.cell(row=row, column=15)
        if sla_met == "✓ TAK":
            o_cell.font = _font(bold=True, color="276221")
        elif sla_met == "-":
            o_cell.font = _font(color="999999")

        # ── Column T — BOTTLENECK colouring ──────────────────────────────────
        t_cell = ws.cell(row=row, column=20)
        if "BOTTLENECK" in str(bottleneck):
            t_cell.fill = _fill("ffc7ce")
            t_cell.font = _font(bold=True, color="9c0006")
        elif "UWAGA" in str(bottleneck):
            t_cell.fill = _fill("ffeb9c")
            t_cell.font = _font(bold=True, color="7B4F00")

        ws.row_dimensions[row].height = 22

    # ── freeze panes at B2 ────────────────────────────────────────────────────
    ws.freeze_panes = "B2"

    # ── data validations extended to row 1000 ────────────────────────────────
    dv_status = DataValidation(
        type="list",
        formula1='"Zakończona,W toku,Brief,Poprawki po WD"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        error="Wybierz wartość z listy.",
        errorTitle="Nieprawidłowy status",
    )
    dv_status.sqref = "D2:D1000"
    ws.add_data_validation(dv_status)

    dv_complex = DataValidation(
        type="list",
        formula1='"Niska,Średnia,Wysoka,Bardzo wysoka,Brief"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        error="Wybierz wartość z listy.",
        errorTitle="Nieprawidłowe skomplikowanie",
    )
    dv_complex.sqref = "H2:H1000"
    ws.add_data_validation(dv_complex)

    # NOTHING below row 16 — sheet ends cleanly here.
    print("  Sheet 1 built — data rows 1-16, rows 17+ empty and ready.")


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — "Dashboard"
# ══════════════════════════════════════════════════════════════════════════════

# Helper: sheet-1 qualified range string
S1 = "'Wdrożenia IdoB'"

def _s1(col_letter, end_row=1000):
    """Return a Sheet1 range reference like 'Wdrożenia IdoB'!D2:D1000"""
    return f"{S1}!{col_letter}2:{col_letter}{end_row}"

def _sep_row(ws, row, text, col_span_end, bg=DK_GREEN, height=28):
    """Merged dark-green separator / section header."""
    merge_ref = f"A{row}:{get_column_letter(col_span_end)}{row}"
    ws.merge_cells(merge_ref)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = _font(bold=True, color=WHITE, size=11)
    c.fill      = _fill(bg)
    c.alignment = _align(h="left")
    ws.row_dimensions[row].height = height
    return c


def build_dashboard(wb: Workbook) -> None:
    ws = wb.create_sheet("Dashboard")

    # default column widths for the stats area
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12

    # chart-data area widths
    for col_letter in ("G","H","J","K","M","N","O","P","R","S","T","V","W","X"):
        ws.column_dimensions[col_letter].width = 16

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION A — PODSUMOWANIE (rows 1-15)
    # ══════════════════════════════════════════════════════════════════════════
    _sep_row(ws, 1, "PODSUMOWANIE WDROZEN", col_span_end=3, height=28)

    # sub-header row 2
    _hdr_cell(ws, 2, 1, "Wskaźnik",  bg=MD_GREEN, h_align="center")
    _hdr_cell(ws, 2, 2, "Wartość",   bg=MD_GREEN, h_align="center")
    ws.row_dimensions[2].height = 20

    SUMMARY = [
        # (row, label, formula_B, num_fmt)
        (3,  "Projekty ogółem",
             f"=COUNTA({_s1('B')})", None),
        (4,  "Zakończone",
             f'=COUNTIF({_s1("D")},"Zakończona")', None),
        (5,  "W toku",
             f'=COUNTIF({_s1("D")},"W toku")', None),
        (6,  "Brief (planowane)",
             f'=COUNTIF({_s1("D")},"Brief")', None),
        (7,  "Poprawki po WD",
             f'=COUNTIF({_s1("D")},"Poprawki po WD")', None),
        (8,  "% ukończenia",
             "=IFERROR(B4/B3,0)", "0.0%"),
        (9,  "Łączny czas budowy (h)",
             f'=SUMIF({_s1("D")},"<>Brief",{_s1("L")})', None),
        (10, "Łączny czas (dni rob. @8h)",
             "=CEILING(B9/8,1)", None),
        (11, "Średni czas budowy (h)",
             f'=IFERROR(ROUND(AVERAGEIF({_s1("D")},"<>Brief",{_s1("L")}),1),0)', None),
        (12, "SLA dotrzymane (%)",
             f'=IFERROR(COUNTIF({_s1("O")},"✓ TAK")/COUNTIF({_s1("O")},"<>-"),0)', "0.0%"),
        (13, "Łączne rundy poprawek",
             f"=SUM({_s1('P')})", None),
        (14, "Projekty z feedbackiem",
             f'=COUNTIF({_s1("S")},"TAK")', None),
        (15, "Bottlenecki",
             f'=COUNTIF({_s1("T")},"*BOTTLENECK*")', None),
    ]

    for row, label, formula, num_fmt in SUMMARY:
        alt = LT_GREEN if row % 2 == 0 else WHITE

        lc = _data_cell(ws, row, 1, label, bg=alt, h_align="left")
        vc = _data_cell(ws, row, 2, formula, bg=alt, bold=True, h_align="center")
        if num_fmt:
            vc.number_format = num_fmt
        ws.row_dimensions[row].height = 20

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION B — PIPELINE (rows 17-22)
    # ══════════════════════════════════════════════════════════════════════════
    ws.column_dimensions["D"].width = 32   # widen for progress bar

    _sep_row(ws, 17, "PIPELINE - STATUS PROJEKTOW", col_span_end=4, height=26)

    for col, hdr in [(1,"Status"),(2,"Liczba"),(3,"%"),(4,"Pasek postępu")]:
        _hdr_cell(ws, 18, col, hdr, bg=MD_GREEN)
    ws.row_dimensions[18].height = 20

    PIPELINE = [
        (19, "Zakończona",     f'=COUNTIF({_s1("D")},"Zakończona")',     *S_ZAKONCZONA),
        (20, "W toku",         f'=COUNTIF({_s1("D")},"W toku")',         *S_W_TOKU),
        (21, "Brief",          f'=COUNTIF({_s1("D")},"Brief")',          *S_BRIEF),
        (22, "Poprawki po WD", f'=COUNTIF({_s1("D")},"Poprawki po WD")', *S_POPRAWKI),
    ]

    for row, label, count_f, bg, fg in PIPELINE:
        # Col A — label
        c1 = _data_cell(ws, row, 1, label, bg=bg, fg=fg, bold=True, h_align="left")
        # Col B — count formula
        c2 = _data_cell(ws, row, 2, count_f, bg=bg, fg=fg, h_align="center")
        # Col C — share %
        c3 = _data_cell(ws, row, 3, f"=IFERROR(B{row}/B3,0)", bg=bg, h_align="center")
        c3.number_format = "0%"
        # Col D — text progress bar
        bar = (
            f'=REPT("█",ROUND(C{row}*20,0))'
            f'&REPT("░",20-ROUND(C{row}*20,0))'
            f'&"  "&TEXT(C{row},"0%")'
        )
        c4 = _data_cell(ws, row, 4, bar, bg=bg, fg=fg,
                        h_align="left", font_name="Courier New")
        ws.row_dimensions[row].height = 20

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION C — CEL ROCZNY (rows 25-28)
    # ══════════════════════════════════════════════════════════════════════════
    _sep_row(ws, 25, "CEL ROCZNY 2026", col_span_end=5, height=26)

    for col, hdr in [(1,"Cel"),(2,"Plan"),(3,"Wykonano"),(4,"% realizacji"),(5,"Pozostało")]:
        _hdr_cell(ws, 26, col, hdr, bg=MD_GREEN)
    ws.row_dimensions[26].height = 20

    goal_rows = [
        (27, "Wdrożenia zakończone", 20,
             f'=COUNTIF({_s1("D")},"Zakończona")',
             "=IFERROR(C27/B27,0)", "=B27-C27", LT_GREEN),
        (28, "Pipeline (wszystkie)", 20,
             f"=COUNTA({_s1('B')})",
             "=IFERROR(C28/B28,0)", "=B28-C28", WHITE),
    ]

    for row, lbl, plan, exec_f, pct_f, rem_f, bg in goal_rows:
        _data_cell(ws, row, 1, lbl,    bg=bg, h_align="left")
        _data_cell(ws, row, 2, plan,   bg=bg, h_align="center", bold=True)
        _data_cell(ws, row, 3, exec_f, bg=bg, h_align="center")
        c4 = _data_cell(ws, row, 4, pct_f, bg=bg, h_align="center")
        c4.number_format = "0%"
        _data_cell(ws, row, 5, rem_f, bg=bg, h_align="center")
        ws.row_dimensions[row].height = 20

    # ══════════════════════════════════════════════════════════════════════════
    # CHART DATA TABLES (cols G onwards, rows 1-20)
    # All formulas in chart tables also use :1000 ranges.
    # ══════════════════════════════════════════════════════════════════════════

    # -- Table 1: Status (G:H, rows 1-5) --
    _hdr_cell(ws, 1, 7, "Status",  bg=DK_GREEN)
    _hdr_cell(ws, 1, 8, "Liczba",  bg=DK_GREEN)
    t1 = [
        ("Zakończona",     f'=COUNTIF({_s1("D")},"Zakończona")'),
        ("W toku",         f'=COUNTIF({_s1("D")},"W toku")'),
        ("Brief",          f'=COUNTIF({_s1("D")},"Brief")'),
        ("Poprawki po WD", f'=COUNTIF({_s1("D")},"Poprawki po WD")'),
    ]
    for i, (lbl, f) in enumerate(t1, start=2):
        _data_cell(ws, i, 7, lbl)
        _data_cell(ws, i, 8, f, h_align="center")

    # -- Table 2: Złożoność (J:K, rows 1-5) --
    _hdr_cell(ws, 1, 10, "Złożoność", bg=DK_GREEN)
    _hdr_cell(ws, 1, 11, "Liczba",    bg=DK_GREEN)
    t2 = [
        ("Niska",         f'=COUNTIF({_s1("H")},"Niska")'),
        ("Średnia",       f'=COUNTIF({_s1("H")},"Średnia")'),
        ("Wysoka",        f'=COUNTIF({_s1("H")},"Wysoka")'),
        ("Bardzo wysoka", f'=COUNTIF({_s1("H")},"Bardzo wysoka")'),
    ]
    for i, (lbl, f) in enumerate(t2, start=2):
        _data_cell(ws, i, 10, lbl)
        _data_cell(ws, i, 11, f, h_align="center")

    # -- Table 3: SLA vs Czas (M:P, rows 1-13) -- static values --
    for col_i, hdr in enumerate(["Projekt","SLA (h)","Czas rzecz. (h)","Delta (h)"], start=13):
        _hdr_cell(ws, 1, col_i, hdr, bg=DK_GREEN)
    t3 = [
        ("CityApart Szczecin", 56,  8,  48),
        ("WawaBed",            80,  20, 60),
        ("Willa Kapitańska",   112, 30, 82),
        ("GoldenApartments",   56,  15, 41),
        ("MW Consulting",      56,  10, 46),
        ("GeoStay",            80,  14, 66),
        ("Mazurski Chill",     56,  12, 44),
        ("Madera Centrum",     80,  20, 60),
        ("Najmar",             112, 35, 77),
        ("EcoCamping",         112, 28, 84),
        ("WCA Wrocław",        24,  5,  19),
        ("SORS Eagle Tower",   112, 22, 90),
    ]
    for i, row_data in enumerate(t3, start=2):
        for col_i, val in enumerate(row_data, start=13):
            h = "left" if isinstance(val, str) else "center"
            _data_cell(ws, i, col_i, val, h_align=h)

    # -- Table 4: Rundy poprawek (R:T, rows 1-13) -- static values --
    for col_i, hdr in enumerate(["Projekt","Rundy","Limit (2)"], start=18):
        _hdr_cell(ws, 1, col_i, hdr, bg=DK_GREEN)
    t4 = [
        ("CityApart Szczecin", 0,  2),
        ("WawaBed",            0,  2),
        ("Willa Kapitańska",   0,  2),
        ("GoldenApartments",   2,  2),
        ("MW Consulting",      0,  2),
        ("GeoStay",            0,  2),
        ("Mazurski Chill",     1,  2),
        ("Madera Centrum",     3,  2),
        ("Najmar",             14, 2),
        ("EcoCamping",         2,  2),
        ("WCA Wrocław",        0,  2),
        ("SORS Eagle Tower",   2,  2),
    ]
    for i, row_data in enumerate(t4, start=2):
        for col_i, val in enumerate(row_data, start=18):
            h = "left" if isinstance(val, str) else "center"
            _data_cell(ws, i, col_i, val, h_align=h)

    # -- Table 5: Trend (V:X, rows 1-8) -- static values --
    for col_i, hdr in enumerate(["Miesiąc","Zakończone","Skumulowane"], start=22):
        _hdr_cell(ws, 1, col_i, hdr, bg=DK_GREEN)
    t5 = [
        ("Styczeń 2026",           0, 0),
        ("Luty 2026",              2, 2),
        ("Marzec 2026",            5, 7),
        ("Kwiecień 2026",          2, 9),
        ("Maj 2026 (plan)",        3, 12),
        ("Czerwiec 2026 (plan)",   3, 15),
        ("Lipiec-Grudzień (plan)", 5, 20),
    ]
    for i, row_data in enumerate(t5, start=2):
        for col_i, val in enumerate(row_data, start=22):
            h = "left" if isinstance(val, str) else "center"
            _data_cell(ws, i, col_i, val, h_align=h)

    # ══════════════════════════════════════════════════════════════════════════
    # CHARTS — all referencing Dashboard sheet data tables
    # ══════════════════════════════════════════════════════════════════════════

    # -- Chart 1: Doughnut "Status wdrożeń" --
    doughnut = DoughnutChart()
    doughnut.title  = "Status wdrozen"
    doughnut.style  = 10
    doughnut.hole   = 50
    doughnut.width  = 13
    doughnut.height = 10

    d_cats = Reference(ws, min_col=7,  min_row=2, max_row=5)
    d_data = Reference(ws, min_col=8,  min_row=1, max_row=5)
    doughnut.add_data(d_data, titles_from_data=True)
    doughnut.set_categories(d_cats)

    slice_colors = ["70ad47", "5b9bd5", "ffd966", "ff7171"]
    if doughnut.series:
        for idx, hex_c in enumerate(slice_colors):
            pt = DataPoint(idx=idx)
            pt.graphicalProperties.solidFill = hex_c
            doughnut.series[0].dPt.append(pt)

    ws.add_chart(doughnut, "A33")   # row 33, col A

    # -- Chart 2: Pie "Złożoność projektów" --
    pie = PieChart()
    pie.title  = "Zlozonosc projektow"
    pie.style  = 10
    pie.width  = 13
    pie.height = 10

    p_cats = Reference(ws, min_col=10, min_row=2, max_row=5)
    p_data = Reference(ws, min_col=11, min_row=1, max_row=5)
    pie.add_data(p_data, titles_from_data=True)
    pie.set_categories(p_cats)

    ws.add_chart(pie, "G33")   # row 33, col G

    # -- Chart 3: BarChart "SLA vs Czas" --
    bar_sla = BarChart()
    bar_sla.type        = "col"
    bar_sla.grouping    = "clustered"
    bar_sla.title       = "SLA vs Czas rzeczywisty (h)"
    bar_sla.style       = 10
    bar_sla.x_axis.title = "Projekt"
    bar_sla.y_axis.title = "Godziny"
    bar_sla.width  = 22
    bar_sla.height = 12

    sla_cats = Reference(ws, min_col=13, min_row=2, max_row=13)
    sla_ref  = Reference(ws, min_col=14, min_row=1, max_row=13)
    real_ref = Reference(ws, min_col=15, min_row=1, max_row=13)
    bar_sla.add_data(sla_ref,  titles_from_data=True)
    bar_sla.add_data(real_ref, titles_from_data=True)
    bar_sla.set_categories(sla_cats)

    if len(bar_sla.series) >= 2:
        bar_sla.series[0].graphicalProperties.solidFill = "70ad47"
        bar_sla.series[1].graphicalProperties.solidFill = "5b9bd5"

    ws.add_chart(bar_sla, "A51")   # row 51, col A

    # -- Chart 4: BarChart "Rundy poprawek" --
    bar_r = BarChart()
    bar_r.type        = "col"
    bar_r.grouping    = "clustered"
    bar_r.title       = "Rundy poprawek per projekt"
    bar_r.style       = 10
    bar_r.x_axis.title = "Projekt"
    bar_r.y_axis.title = "Rundy"
    bar_r.width  = 17
    bar_r.height = 12

    r_cats  = Reference(ws, min_col=18, min_row=2, max_row=13)
    r_ref   = Reference(ws, min_col=19, min_row=1, max_row=13)
    lim_ref = Reference(ws, min_col=20, min_row=1, max_row=13)
    bar_r.add_data(r_ref,   titles_from_data=True)
    bar_r.add_data(lim_ref, titles_from_data=True)
    bar_r.set_categories(r_cats)

    if len(bar_r.series) >= 2:
        bar_r.series[0].graphicalProperties.solidFill = "ff7171"
        bar_r.series[1].graphicalProperties.solidFill = "ffd966"

    ws.add_chart(bar_r, "J51")   # row 51, col J

    # -- Chart 5: LineChart "Trend miesięczny" --
    line = LineChart()
    line.title        = "Trend miesieczny - cel: 20 wdrozen 2026"
    line.style        = 10
    line.y_axis.title = "Liczba wdrozen"
    line.x_axis.title = "Miesiac"
    line.grouping     = "standard"
    line.width  = 29
    line.height = 11

    tr_cats  = Reference(ws, min_col=22, min_row=2, max_row=8)
    tr_done  = Reference(ws, min_col=23, min_row=1, max_row=8)
    tr_cum   = Reference(ws, min_col=24, min_row=1, max_row=8)
    line.add_data(tr_done, titles_from_data=True)
    line.add_data(tr_cum,  titles_from_data=True)
    line.set_categories(tr_cats)

    if len(line.series) >= 2:
        line.series[0].graphicalProperties.line.solidFill = "70ad47"
        line.series[0].graphicalProperties.line.width     = 20000
        line.series[1].graphicalProperties.line.solidFill = "5b9bd5"
        line.series[1].graphicalProperties.line.width     = 20000

    ws.add_chart(line, "A69")   # row 69, col A

    print("  Sheet 2 (Dashboard) built — 3 stat sections + 5 data tables + 5 charts.")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    wb = Workbook()

    print("Building Sheet 1 (Wdrożenia IdoB) ...")
    build_sheet1(wb)

    print("Building Sheet 2 (Dashboard) ...")
    build_dashboard(wb)

    wb.save(OUTPUT_PATH)

    fsize = os.path.getsize(OUTPUT_PATH)
    print()
    print("File saved successfully:")
    print(f"  {OUTPUT_PATH}")
    print(f"  Size: {fsize:,} bytes ({fsize/1024:.1f} KB)")
    print(f"  openpyxl: {openpyxl.__version__}")
    print()
    print("Architecture:")
    print("  Sheet 1 — data rows 1-16 only, rows 17+ clean for new projects")
    print("  Sheet 2 — Dashboard with all stats + 5 charts (formulas use :1000 ranges)")


if __name__ == "__main__":
    main()
